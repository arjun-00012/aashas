import json
import razorpay
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from django.conf import settings
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .models import Category, Product, Order, OrderItem, Profile, ContactMessage
from .forms import RegistrationForm, ProfileUpdateForm, CategoryForm, ProductForm

razorpay_client = razorpay.Client(auth=(settings.RAZORPAY_KEY_ID, settings.RAZORPAY_KEY_SECRET))

def cart_context_processor(request):
    cart = request.session.get('cart', {})
    total_qty = sum(cart.values()) if cart else 0
    return {
        'cart_item_count': total_qty,
        'all_categories': Category.objects.all(),
    }

def home(request):
    categories = Category.objects.prefetch_related('products').all()
    return render(request, 'index.html', {'categories': categories})

# --- Auth Views ---
def register_view(request):
    if request.method == 'POST':
        form = RegistrationForm(request.POST)
        if form.is_valid():
            user = User.objects.create_user(
                username=form.cleaned_data['username'],
                password=form.cleaned_data['password']
            )
            profile = user.profile
            profile.phone_number = form.cleaned_data['phone_number']
            profile.save()
            login(request, user)
            return redirect('home')
    else:
        form = RegistrationForm()
    return render(request, 'register.html', {'form': form})

def login_view(request):
    if request.method == 'POST':
        u = request.POST.get('username')
        p = request.POST.get('password')
        user = authenticate(request, username=u, password=p)
        if user:
            login(request, user)
            next_url = request.GET.get('next', 'home')
            return redirect(next_url)
        return render(request, 'login.html', {'error': 'Invalid Username or Password.'})
    return render(request, 'login.html')

def logout_view(request):
    logout(request)
    return redirect('home')

@login_required
def profile_view(request):
    profile, _ = Profile.objects.get_or_create(user=request.user)
    if request.method == 'POST':
        form = ProfileUpdateForm(request.POST, instance=profile)
        if form.is_valid():
            form.save()
            return redirect('profile')
    else:
        form = ProfileUpdateForm(instance=profile)
    orders = Order.objects.filter(user=request.user).order_by('-created_at')
    return render(request, 'profile.html', {'form': form, 'orders': orders, 'profile': profile})

# --- Cart Views ---
def add_to_cart(request, product_id):
    cart = request.session.get('cart', {})
    pid = str(product_id)
    cart[pid] = cart.get(pid, 0) + 1
    request.session['cart'] = cart
    return redirect(request.META.get('HTTP_REFERER', 'home'))

def cart_view(request):
    cart = request.session.get('cart', {})
    cart_items = []
    total = 0
    stale_pids = []
    for pid, qty in list(cart.items()):
        product = Product.objects.filter(id=pid).first()
        if not product:
            stale_pids.append(pid)
            continue
        subtotal = float(product.current_price) * qty
        total += subtotal
        cart_items.append({'product': product, 'quantity': qty, 'subtotal': subtotal})
    
    if stale_pids:
        for pid in stale_pids:
            cart.pop(pid, None)
        request.session['cart'] = cart

    return render(request, 'cart.html', {'cart_items': cart_items, 'total': total})

def update_cart(request, product_id, action):
    cart = request.session.get('cart', {})
    pid = str(product_id)
    if pid in cart:
        if action == 'increase':
            cart[pid] += 1
        elif action == 'decrease':
            cart[pid] -= 1
            if cart[pid] <= 0:
                del cart[pid]
        elif action == 'remove':
            del cart[pid]
    request.session['cart'] = cart
    return redirect('cart')

# --- Razorpay Checkout ---
def checkout_view(request):
    cart = request.session.get('cart', {})
    if not cart:
        return redirect('home')
    
    valid_items = {}
    total = 0
    stale_pids = []
    for pid, qty in list(cart.items()):
        product = Product.objects.filter(id=pid).first()
        if product:
            valid_items[pid] = (product, qty)
            total += float(product.current_price) * qty
        else:
            stale_pids.append(pid)

    if stale_pids:
        for pid in stale_pids:
            cart.pop(pid, None)
        request.session['cart'] = cart

    if not valid_items:
        return redirect('home')
    
    default_address = ''
    default_phone = ''
    default_name = ''
    if request.user.is_authenticated:
        profile, _ = Profile.objects.get_or_create(user=request.user)
        default_address = profile.address or ''
        default_phone = profile.phone_number or ''
        default_name = request.user.username

    if request.method == 'POST':
        full_name = request.POST.get('full_name')
        phone = request.POST.get('phone_number')
        address = request.POST.get('shipping_address')

        rzp_order = razorpay_client.order.create({
            'amount': int(total * 100),
            'currency': 'INR',
            'payment_capture': '1'
        })

        order = Order.objects.create(
            user=request.user if request.user.is_authenticated else None,
            full_name=full_name,
            phone_number=phone,
            shipping_address=address,
            total_price=total,
            razorpay_order_id=rzp_order['id'],
            payment_status='Pending'
        )

        for pid, (product, qty) in valid_items.items():
            OrderItem.objects.create(
                order=order,
                product=product,
                price=product.current_price,
                quantity=qty
            )

        return JsonResponse({
            'razorpay_key': settings.RAZORPAY_KEY_ID,
            'amount': rzp_order['amount'],
            'razorpay_order_id': rzp_order['id'],
            'db_order_id': order.id
        })

    return render(request, 'checkout.html', {
        'total': total,
        'default_address': default_address,
        'default_phone': default_phone,
        'default_name': default_name
    })

@csrf_exempt
def payment_verify(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        order_id = data.get('db_order_id')
        params = {
            'razorpay_order_id': data.get('razorpay_order_id'),
            'razorpay_payment_id': data.get('razorpay_payment_id'),
            'razorpay_signature': data.get('razorpay_signature')
        }
        try:
            razorpay_client.utility.verify_payment_signature(params)
            order = Order.objects.get(id=order_id)
            order.payment_status = 'Completed'
            order.razorpay_payment_id = data.get('razorpay_payment_id')
            order.save()
            request.session['cart'] = {}
            return JsonResponse({'status': 'success'})
        except Exception:
            return JsonResponse({'status': 'failed'}, status=400)

def contact_submit(request):
    if request.method == 'POST':
        ContactMessage.objects.create(
            name=request.POST.get('name'),
            email=request.POST.get('email'),
            subject=request.POST.get('subject'),
            message=request.POST.get('message')
        )
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'status': 'success'})
        messages.success(request, "Your message has been sent successfully!")
        return redirect(request.META.get('HTTP_REFERER', 'home'))
    return JsonResponse({'status': 'invalid'}, status=400)

# --- AdminPP Custom Dashboard ---
def staff_required(view_func):
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated or not request.user.is_staff:
            return redirect('/login/?next=' + request.path)
        return view_func(request, *args, **kwargs)
    return wrapper

@staff_required
def adminpp_dashboard(request):
    return render(request, 'adminpp_dashboard.html', {
        'categories': Category.objects.all(),
        'products': Product.objects.all(),
        'inquiries': ContactMessage.objects.all().order_by('-created_at')
    })

@staff_required
def adminpp_orders(request):
    category_filter = request.GET.get('category', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    export_excel = request.GET.get('export', '')

    orders = Order.objects.filter(payment_status='Completed').prefetch_related('items__product__category').order_by('-created_at')

    if category_filter:
        orders = orders.filter(items__product__category__name__iexact=category_filter).distinct()
    if start_date:
        orders = orders.filter(created_at__date__gte=start_date)
    if end_date:
        orders = orders.filter(created_at__date__lte=end_date)

    if export_excel == 'true':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Orders Data"

        headers = ['Order ID', 'Customer Name', 'Phone', 'Address', 'Items Purchased', 'Categories', 'Total Price', 'Payment ID', 'Date']
        ws.append(headers)

        header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
        header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")

        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for o in orders:
            items_str = ", ".join([f"{i.product.name} (x{i.quantity})" for i in o.items.all()])
            cats_str = ", ".join(list(set([i.product.category.name for i in o.items.all()])))
            ws.append([
                o.id,
                o.full_name,
                o.phone_number,
                o.shipping_address,
                items_str,
                cats_str,
                float(o.total_price),
                o.razorpay_payment_id or '',
                o.created_at.strftime('%Y-%m-%d %H:%M')
            ])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="aashas_orders_export.xlsx"'
        wb.save(response)
        return response

    return render(request, 'adminpp_orders.html', {
        'orders': orders,
        'categories': Category.objects.all(),
        'selected_category': category_filter,
        'start_date': start_date,
        'end_date': end_date
    })

@staff_required
def category_create_or_edit(request, pk=None):
    category = get_object_or_404(Category, pk=pk) if pk else None
    form = CategoryForm(request.POST or None, request.FILES or None, instance=category)
    if request.method == 'POST':
        if form.is_valid():
            cat = form.save()
            action_text = "updated" if pk else "created"
            messages.success(request, f'Category "{cat.name}" has been {action_text} successfully!')
            return redirect('adminpp_dashboard')
        else:
            messages.error(request, 'Please correct the form errors below.')
    return render(request, 'category_form.html', {'form': form, 'category': category})

@staff_required
def category_delete(request, pk):
    category = get_object_or_404(Category, pk=pk)
    cat_name = category.name
    category.delete()
    messages.success(request, f'Category "{cat_name}" deleted successfully.')
    return redirect('adminpp_dashboard')

@staff_required
def product_create_or_edit(request, pk=None):
    product = get_object_or_404(Product, pk=pk) if pk else None
    form = ProductForm(request.POST or None, request.FILES or None, instance=product)
    if request.method == 'POST':
        if form.is_valid():
            p = form.save()
            action_text = "updated" if pk else "created"
            messages.success(request, f'Product "{p.name}" has been {action_text} successfully!')
            return redirect('adminpp_dashboard')
        else:
            messages.error(request, 'Please correct the form errors below.')
    return render(request, 'product_form.html', {'form': form, 'product': product})

@staff_required
def product_delete(request, pk):
    product = get_object_or_404(Product, pk=pk)
    p_name = product.name
    product.delete()
    messages.success(request, f'Product "{p_name}" deleted successfully.')
    return redirect('adminpp_dashboard')